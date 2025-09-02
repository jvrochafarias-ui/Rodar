import streamlit as st
import pandas as pd
from datetime import datetime, timedelta
import unicodedata
import io
import base64
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import random

# ------------------------
# Funções auxiliares
# ------------------------
def normalizar_coluna(col):
    col = str(col).strip().upper()
    col = unicodedata.normalize('NFKD', col).encode('ASCII', 'ignore').decode('ASCII')
    col = col.replace(" ", "_")
    return col

def normalizar_texto(txt):
    if pd.isna(txt):
        return ""
    txt = str(txt).strip().upper()
    txt = unicodedata.normalize('NFKD', txt).encode('ASCII', 'ignore').decode('ASCII')
    return txt

def esta_disponivel(row, data):
    """
    Verifica se a pessoa está disponível na data desejada.
    Bloqueia qualquer pessoa que tenha INICIO e FIM de indisponibilidade que inclua a data.
    """
    if pd.isna(data):
        return True

    data = pd.to_datetime(data).normalize()  # normaliza a data para comparar

    inicio = row.get('INICIO_INDISPONIBILIDADE', pd.NaT)
    fim = row.get('FIM_INDISPONIBILIDADE', pd.NaT)

    # Se a pessoa marcou "SIM" na coluna de indisponibilidade geral
    if str(inicio).strip().upper() == 'SIM':
        return False

    # Tenta converter as colunas para datas (se forem strings)
    try:
        if pd.notna(inicio):
            inicio = pd.to_datetime(inicio, dayfirst=True).normalize()
        if pd.notna(fim):
            fim = pd.to_datetime(fim, dayfirst=True).normalize()
    except Exception:
        return True  # Se não for data válida, ignora

    # Bloqueia qualquer data dentro do intervalo de indisponibilidade
    if pd.notna(inicio) and pd.notna(fim):
        if inicio <= data <= fim:
            return False

    return True

# ------------------------
# Processamento da distribuição
# ------------------------
def processar_distribuicao(arquivo_excel):
    xls = pd.ExcelFile(arquivo_excel)
    sheet_name = 'Planilha1' if 'Planilha1' in xls.sheet_names else xls.sheet_names[0]
    df = pd.read_excel(xls, sheet_name=sheet_name)

    # Normalizar colunas
    df.columns = [normalizar_coluna(col) for col in df.columns]

    # Garantir coluna NOME
    colunas_possiveis_nome = ['NOME', 'NOME_COMPLETO', 'NOME_PESSOA']
    for col in colunas_possiveis_nome:
        if col in df.columns:
            df['NOME'] = df[col]
            break
    if 'NOME' not in df.columns:
        st.error(f"❌ Erro: não foi possível localizar a coluna de nomes. Colunas disponíveis: {df.columns.tolist()}")
        return None, pd.DataFrame(), pd.DataFrame(), io.BytesIO()

    # Criar colunas padrão
    df['INDISPONIBILIDADE'] = df.get('INDISPONIBILIDADE', pd.Series("NAO")).fillna("NAO")
    df['PRESIDENTE_DE_BANCA'] = df.get('PRESIDENTE_DE_BANCA', pd.Series("NAO")).fillna("NAO")
    df['MUNICIPIO_ORIGEM'] = df.get('MUNICIPIO_ORIGEM', pd.Series("")).fillna("")
    df['INICIO_INDISPONIBILIDADE'] = df.get('INICIO_INDISPONIBILIDADE', pd.NaT)
    df['FIM_INDISPONIBILIDADE'] = df.get('FIM_INDISPONIBILIDADE', pd.NaT)

    distribuicoes = []                  
    pessoas_agendadas = {}
    contador_convocacoes = {nome: 0 for nome in df['NOME'].unique()}
    historico_municipio = {nome: [] for nome in df['NOME'].unique()}
    presidentes_ja_convocados = set()

    dias_distribuicao = df[['DIA', 'DATA', 'MUNICIPIO', 'CATEGORIA', 'QUANTIDADE']].dropna(subset=['DIA'])
    candidatos_df = df[['NOME', 'CATEGORIA', 'INDISPONIBILIDADE', 'PRESIDENTE_DE_BANCA',
                        'MUNICIPIO_ORIGEM', 'INICIO_INDISPONIBILIDADE', 'FIM_INDISPONIBILIDADE']].dropna(subset=['NOME'])

    traducao_dias_eng = {'MONDAY':'SEGUNDA','TUESDAY':'TERCA','WEDNESDAY':'QUARTA','THURSDAY':'QUINTA','FRIDAY':'SEXTA'}

    for _, row in dias_distribuicao.iterrows():
        dia_raw = str(row['DIA']).strip().upper()
        municipio = row['MUNICIPIO']
        categorias_necessarias = [cat.strip() for cat in str(row['CATEGORIA']).split(',')]
        try:
            quantidade = int(row['QUANTIDADE'])
        except ValueError:
            continue

        data_municipio = pd.to_datetime(row.get('DATA', row['DIA']), dayfirst=True, errors='coerce')
        dia_semana_pt = traducao_dias_eng.get(data_municipio.strftime('%A').upper(), dia_raw) if pd.notna(data_municipio) else dia_raw

        candidatos = candidatos_df[
            (candidatos_df['CATEGORIA'].apply(lambda x: any(cat in str(x) for cat in categorias_necessarias))) &
            (candidatos_df['MUNICIPIO_ORIGEM'].apply(normalizar_texto) != normalizar_texto(municipio))
        ].copy()
        # Aplica indisponibilidade, incluindo o período de datas
        candidatos = candidatos[candidatos.apply(lambda x: esta_disponivel(x, data_municipio), axis=1)]

        if dia_semana_pt in pessoas_agendadas:
            candidatos = candidatos[~candidatos['NOME'].isin(pessoas_agendadas[dia_semana_pt])]

        dia_anterior = data_municipio - timedelta(days=1) if pd.notna(data_municipio) else None
        if dia_anterior:
            def nao_consecutivo_mesmo_municipio(nome):
                for mun, data_mun in historico_municipio.get(nome, []):
                    if normalizar_texto(mun) == normalizar_texto(municipio) and data_mun == dia_anterior:
                        return False
                return True
            candidatos = candidatos[candidatos['NOME'].apply(nao_consecutivo_mesmo_municipio)]

        if pd.notna(data_municipio):
            week_start = data_municipio - timedelta(days=data_municipio.weekday())
            week_end = week_start + timedelta(days=6)
            def nao_mesmo_municipio_semana(nome):
                for mun, data_mun in historico_municipio.get(nome, []):
                    if normalizar_texto(mun) == normalizar_texto(municipio) and week_start <= data_mun <= week_end:
                        return False
                return True
            candidatos = candidatos[candidatos['NOME'].apply(nao_mesmo_municipio_semana)]

        if candidatos.empty:
            continue

        candidatos_list = candidatos.assign(
            CONVOCACOES=candidatos['NOME'].map(contador_convocacoes)
        ).sort_values(by="CONVOCACOES", ascending=True)

        candidatos_list = candidatos_list.groupby("CONVOCACOES", group_keys=False).apply(
            lambda x: x.sample(frac=1, random_state=random.randint(0, 10000))
        ).reset_index(drop=True)

        selecionados = candidatos_list.head(min(quantidade, len(candidatos_list)))
        if selecionados.empty:
            continue

        presidentes = selecionados[selecionados['PRESIDENTE_DE_BANCA'].str.upper() == 'SIM']
        presidente_nome = None
        for p in presidentes['NOME']:
            if p not in presidentes_ja_convocados:
                presidente_nome = p
                break
        if presidente_nome is None and not presidentes.empty:
            presidente_nome = presidentes.iloc[0]['NOME']

        if presidente_nome:
            presidentes_ja_convocados.add(presidente_nome)

        for _, pessoa in selecionados.iterrows():
            distribuicoes.append({
                "DIA": dia_semana_pt,
                "DATA": data_municipio.strftime("%d/%m/%y") if pd.notna(data_municipio) else "",
                "MUNICIPIO": municipio,
                "NOME": pessoa['NOME'],
                "CATEGORIA": pessoa['CATEGORIA'],
                "PRESIDENTE": "SIM" if pessoa['NOME'] == presidente_nome else "NAO"
            })
            contador_convocacoes[pessoa['NOME']] += 1
            historico_municipio[pessoa['NOME']].append((municipio, data_municipio))

        if dia_semana_pt not in pessoas_agendadas:
            pessoas_agendadas[dia_semana_pt] = []
        pessoas_agendadas[dia_semana_pt].extend(selecionados['NOME'].tolist())

    df_convocados = pd.DataFrame(distribuicoes)

    # Lista de não convocados
    nao_convocados_lista = []
    for _, row in dias_distribuicao.iterrows():
        municipio = row['MUNICIPIO']
        categorias_necessarias = [cat.strip() for cat in str(row['CATEGORIA']).split(',')]
        data_municipio = pd.to_datetime(row.get('DATA', row['DIA']), dayfirst=True, errors='coerce')
        dia_semana_pt = traducao_dias_eng.get(
            data_municipio.strftime('%A').upper(),
            str(row['DIA']).upper()
        ) if pd.notna(data_municipio) else str(row['DIA']).upper()

        candidatos = candidatos_df[
            (candidatos_df['CATEGORIA'].apply(lambda x: any(cat in str(x) for cat in categorias_necessarias))) &
            (candidatos_df['MUNICIPIO_ORIGEM'].apply(normalizar_texto) != normalizar_texto(municipio))
        ].copy()
        candidatos = candidatos[candidatos.apply(lambda x: esta_disponivel(x, data_municipio), axis=1)]

        nomes_convocados = df_convocados[
            (df_convocados['DIA'] == dia_semana_pt) &
            (df_convocados['DATA'] == (data_municipio.strftime("%d/%m/%y") if pd.notna(data_municipio) else ""))
        ]['NOME'].tolist()

        for n in candidatos['NOME'].tolist():
            if n not in nomes_convocados:
                nao_convocados_lista.append({"NOME": n, "DIA": dia_semana_pt})

    df_nao_convocados = pd.DataFrame(nao_convocados_lista).drop_duplicates(subset=["NOME", "DIA"])

    # Salva Excel
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Convocados"
    for r_idx, row in enumerate(dataframe_to_rows(df_convocados, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws1.cell(row=r_idx, column=c_idx, value=value)

    ws2 = wb.create_sheet("Nao Convocados")
    for r_idx, row in enumerate(dataframe_to_rows(df_nao_convocados, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws2.cell(row=r_idx, column=c_idx, value=value)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    nome_arquivo_saida = f'distribuicao_{datetime.now().strftime("%B").upper()}.xlsx'
    return nome_arquivo_saida, df_convocados, df_nao_convocados, output

# ------------------------
# Interface Streamlit (layout moderno)
# ------------------------
st.set_page_config(page_title="Distribuição Aleatória", page_icon="📊", layout="centered")

# CSS moderno
page_bg = """
<style>
.stApp {
    background: linear-gradient(135deg, #002b45, #014d63, #028090);
    background-attachment: fixed;
    color: white;
    font-family: 'Segoe UI', sans-serif;
}
.main-card {
    background: rgba(255, 255, 255, 0.08);
    border-radius: 20px;
    padding: 40px;
    box-shadow: 0 8px 25px rgba(0,0,0,0.4);
    text-align: center;
    margin-top: 40px;
}
.main-card h1 {
    font-size: 2.2rem;
    font-weight: 700;
    color: #ffffff;
    margin-bottom: 15px;
}
.main-card p {
    font-size: 1.1rem;
    color: #dcdcdc;
    margin-bottom: 30px;
}
.stButton button {
    background: linear-gradient(90deg, #00c6ff, #0072ff);
    color: white;
    border: none;
    border-radius: 12px;
    padding: 12px 25px;
    font-size: 1rem;
    font-weight: bold;
    transition: 0.3s;
}
.stButton button:hover {
    transform: scale(1.05);
    background: linear-gradient(90deg, #0072ff, #00c6ff);
}
</style>
"""
st.markdown(page_bg, unsafe_allow_html=True)

# Layout inicial
st.markdown(
    """
    <div class="main-card">
        <h1>📊 Distribuição Aleatória de Pessoas</h1>
        <p>Envie sua planilha Excel e gere automaticamente uma distribuição de convocados e não convocados de forma rápida e organizada.</p>
    </div>
    """,
    unsafe_allow_html=True
)

# Upload do arquivo
arquivo = st.file_uploader("📁 Envie a planilha (.xlsx)", type="xlsx")

if arquivo:
    st.markdown("### ⚙️ Processamento")
    st.info("Clique no botão abaixo para gerar a distribuição.")

    if st.button("🔄 Gerar Distribuição"):
        with st.spinner("Processando..."):
            nome_saida, df_convocados, df_nao_convocados, arquivo_excel = processar_distribuicao(arquivo)

        if df_convocados.empty and df_nao_convocados.empty:
            st.error("⚠️ Não foi possível gerar a distribuição. Verifique a planilha enviada.")
        else:
            st.success("✅ Distribuição gerada com sucesso!")

            col1, col2 = st.columns(2)

            with col1:
                st.markdown("### 👥 Convocados")
                st.dataframe(df_convocados, use_container_width=True)

            with col2:
                st.markdown("### 🚫 Não Convocados")
                st.dataframe(df_nao_convocados, use_container_width=True)

            # Download Bonito
            b64 = base64.b64encode(arquivo_excel.read()).decode()
            st.markdown(
                f"""
                <div style="text-align:center; margin-top:30px;">
                    <a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}"
                       download="{nome_saida}"
                       target="_blank"
                       style="background:linear-gradient(90deg, #00c6ff, #0072ff); padding:12px 25px; color:white; text-decoration:none; border-radius:12px; font-size:16px; font-weight:bold;">
                        ⬇️ Baixar Excel
                    </a>
                </div>
                """,
                unsafe_allow_html=True
            )






